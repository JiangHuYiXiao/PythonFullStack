# -*- coding:utf-8 -*-
# @Author         : 江湖一笑
# @Time           : 2021/10/8 15:07
# @Software       : PythonFullStack
# @Python_verison : 3.7

'''
python中操作excel需要用到第三方库openpyxl，这是一个专门操作excel的库。


'''

# 导入工作簿
from openpyxl import load_workbook

# 创建wb对象
wb = load_workbook('case_data.xlsx')
print(wb)

# 获取sheet页
sheet1 = wb['Sheet1']
print(sheet1)

# # 获取sheet页中的某个单元格cell,是一个对象
# cell = sheet1.cell(row=1,column=1)
# print(cell)
#
# # 获取单元格中的值cell.value
# print(cell.value)


# 获取整个sheet页的所有数据值,sheet.values返回7一个generator 对象

# data = sheet1.values
# print(data)


# generator对象可以使用for循环进行取值，或者list进行转换

# for i in data:
#     print(i)

# list将generator转换成一个list包含所有测试数据
data = list(sheet1.values)
print(data)