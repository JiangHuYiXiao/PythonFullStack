# -*- coding:utf-8 -*-
# @Author         : 江湖一笑
# @Time           : 2021/10/8 16:23
# @Software       : PythonFullStack
# @Python_verison : 3.7


'''
函数封装的步骤
1、函数体的逻辑
2、参数
3、返回值
'''

# 导入工作簿
from openpyxl import load_workbook

def read_excel(file,sheet_name):
    '''
    :param file:
    :param sheet_name:
    :return: 把excel中的数据返回成一个列表类型
    '''
    # 创建wb对象
    wb = load_workbook(file)

    # 获取sheet页
    sheet1 = wb[sheet_name]

    # list将generator转换成一个list包含所有测试数据
    data = list(sheet1.values)
    return data



def read_excel1(file,sheet_name):
    '''
    :param file:
    :param sheet_name:
    :return: 把excel中的数据返回成一个列表类型
    '''
    # 创建wb对象
    wb = load_workbook(file)

    # 获取sheet页
    sheet1 = wb[sheet_name]

    # list将generator转换成一个list包含所有测试数据
    data = list(sheet1.values)
    row = data[1:]
    # for i in res:
    #     res_data = dict(zip(data[0],i))
    #     print(res_data)
    rows = [dict(zip(data[0], i)) for i in row]
    return rows


print(read_excel1(file='case_data.xlsx', sheet_name='Sheet1'))
# [{'id': 1, 'title': 'test1', 'data': 'jianghu', 'expected': '成功'},
# {'id': 2, 'title': 'test2', 'data': 'yixiu', 'expected': '失败'},
# {'id': 3, 'title': 'test3', 'data': 'hhhh', 'expected': '成功'}]